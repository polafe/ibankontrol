#kütüphaneler
from selenium import webdriver
import sys
from selenium.webdriver.chrome.options import Options
from  openpyxl import *
import xlsxwriter

#değişkenler
workbook = xlsxwriter.Workbook("Sonuç.xlsx")
worksheet = workbook.add_worksheet(name="ibansorgu")
kitap = load_workbook("dosya.xlsx")
sheet = kitap.active
worksheet.write(0, 0, "İban no",)
worksheet.write(0, 1, "Sonuç")
croption = webdriver.ChromeOptions()
croption.add_argument('--headless')
yol = "chromedriver.exe"
driver = webdriver.Chrome(executable_path=yol, options=croption)
driver.get("https://www.e-iban.com/")
sat = 2

#döngü
for i in range(1, sheet.max_row):
    #okuma
    deger = sheet.cell(row=sat, column=1)
    #yazma
    driver.find_element_by_id("ibangir").clear()
    driver.find_element_by_id("ibangir").send_keys(deger.value)
    driver.find_element_by_xpath("/html/body/div[1]/section/div[8]/form/div/div/button").click()
    sonuc = driver.find_element_by_xpath("/html/body/div[1]/section/div[8]/div[1]/p[2]").text
    if sonuc == "Girdiğiniz iban doğrudur.":
        sonuc = "Doğru"


    else:
        sonuc = "Hatalıdır"


    print(sonuc)
    worksheet.write(sat - 1, 0, deger.value)
    worksheet.write(sat - 1, 1, sonuc)
    sat = sat + 1
    print(deger.value)



print("Program başarıyla tamamlanmıştır sonuç dosyası oluştuysa bu ekranı kapatabilirsiniz")
    
workbook.close()
driver.close()

